import streamlit as st
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from PIL import Image as PILImage
import io

st.set_page_config(page_title="점포 입회 점검 시스템", layout="centered")
st.title("📄 세로형 입회 점검 리포트")

# 1. 입력 양식
with st.form("inspection_form"):
    visit_date = st.date_input("📅 입회날짜")
    business_type = st.selectbox("🏬 업태", ["GMS", "SM", "APP", "기타"])
    company_name = st.text_input("🏢 기업명")
    store_name = st.text_input("🏪 점포명")
    sv_name = st.text_input("👤 SV명")
    visit_purpose = st.selectbox("🎯 입회목적", ["SV입회", "신규기업", "현장지도", "기타"])
    visit_result = st.text_area("📝 입회결과")
    uploaded_files = st.file_uploader("📸 사진 업로드 (여러 장 가능)", type=["jpg", "jpeg", "png"], accept_multiple_files=True)
    
    submit_button = st.form_submit_button("✨ 엑셀 리포트 생성")

# 2. 실행 로직
if submit_button:
    if not (store_name and company_name):
        st.warning("⚠️ 기업명과 점포명을 입력해 주세요.")
        st.stop()

    try:
        with st.spinner('리포트를 생성 중입니다...'):
            wb = Workbook()
            ws = wb.active
            ws.title = "입회리포트"

            # 디자인 스타일
            label_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            bold_font = Font(bold=True)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

            content = [
                ("입회날짜", str(visit_date)),
                ("업태", business_type),
                ("기업명", company_name),
                ("점포명", store_name),
                ("SV명", sv_name),
                ("입회목적", visit_purpose),
                ("입회결과", visit_result)
            ]

            # 텍스트 데이터 작성
            for i, (label, value) in enumerate(content, 1):
                ws.cell(row=i, column=1).value = label
                ws.cell(row=i, column=1).fill = label_fill
                ws.cell(row=i, column=1).font = bold_font
                ws.cell(row=i, column=1).alignment = center_align
                ws.cell(row=i, column=1).border = border
                
                ws.cell(row=i, column=2).value = value
                ws.cell(row=i, column=2).alignment = left_align
                ws.cell(row=i, column=2).border = border
                ws.row_dimensions[i].height = 30 if label != "입회결과" else 120

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 60

            # 📸 사진 세로로 추가하기 (질문하신 잘린 뒷부분)
            if uploaded_files:
                current_row = len(content) + 1
                for uploaded_file in uploaded_files:
                    try:
                        # 항목 라벨 표시
                        ws.cell(row=current_row, column=1).value = "현장사진"
                        ws.cell(row=current_row, column=1).fill = label_fill
                        ws.cell(row=current_row, column=1).border = border
                        
                        # 이미지 처리
                        img = PILImage.open(uploaded_file)
                        img = img.convert("RGB") # 포맷 통일
                        img.thumbnail((400, 400)) # 크기 조절
                        
                        img_byte_arr = io.BytesIO()
                        img.save(img_byte_arr, format='JPEG', quality=75) # 용량 최적화
                        img_byte_arr.seek(0)
                        
                        img_for_excel = OpenpyxlImage(img_byte_arr)
                        ws.add_image(img_for_excel, f'B{current_row}')
                        
                        # 사진 칸 높이 설정 및 테두리
                        ws.row_dimensions[current_row].height = 250
                        ws.cell(row=current_row, column=2).border = border
                        current_row += 1
                    except Exception as img_e:
                        st.error(f"사진 처리 중 오류: {img_e}")

            # 3. 엑셀 파일 저장 및 다운로드 버튼 생성 (중요!)
            excel_data = io.BytesIO()
            wb.save(excel_data)
            excel_data.seek(0)
            
            st.success("✅ 리포트가 완성되었습니다!")
            st.download_button(
                label="📁 엑셀 파일 다운로드",
                data=excel_data,
                file_name=f"입회보고서_{store_name}_{visit_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    except Exception as e:
        st.error(f"🚫 에러 발생: {e}")
